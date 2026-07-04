"""
Management command: normalizar_turnos

Normaliza la hoja `BASE` de "TURNOS LIVE OPS.xlsx" (grilla de franjas 0/1) a turnos
legibles entrada/salida por agente y día, para un mes dado. Opcionalmente los carga
en el modelo TurnoEquipo.

    python manage.py normalizar_turnos 2026-06            # solo muestra
    python manage.py normalizar_turnos 2026-06 --cargar   # además inserta en BD
"""

from datetime import date, datetime, time, timedelta
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand, CommandError

from apps.liveops.services import guardar_turnos, normalizar_trabajador
from core.conf import settings as env
from core.horarios import DIAS_ORDEN, get_semana_inicio

# Índices de columnas en la hoja BASE (0-based).
COL_YEAR, COL_DIA, COL_FECHA, COL_AGENTE = 2, 6, 8, 10
COL_GRID_INI, COL_GRID_FIN = 16, 58  # franjas de 30 min (07:00 → 03:30)


def _excel_por_defecto() -> Path:
    return Path(env.ALL_IN_ONE_DATA).parent / "TURNOS LIVE OPS.xlsx"


def _salida(t: time) -> time:
    return (datetime.combine(date(2000, 1, 1), t) + timedelta(minutes=30)).time()


class Command(BaseCommand):
    """Convierte la grilla 0/1 de la hoja BASE en turnos y, con --cargar, los inserta."""

    help = "Normaliza turnos del Excel de LiveOps (hoja BASE) para un mes (AAAA-MM)."

    def add_arguments(self, parser):
        parser.add_argument("mes", help="Mes a normalizar en formato AAAA-MM (ej: 2026-06).")
        parser.add_argument("--excel", default=str(_excel_por_defecto()))
        parser.add_argument("--cargar", action="store_true", help="Inserta en TurnoEquipo.")

    def handle(self, *args, **opts):
        try:
            anio, mes = (int(x) for x in opts["mes"].split("-"))
        except ValueError as exc:
            raise CommandError("Formato de mes inválido; usa AAAA-MM.") from exc

        ruta = Path(opts["excel"])
        if not ruta.exists():
            raise CommandError(f"No existe el Excel: {ruta}")

        wb = openpyxl.load_workbook(str(ruta), read_only=True, data_only=True)
        ws = wb["BASE"]
        filas_xl = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=62, values_only=True))
        hdr = filas_xl[0]
        grid = [(c, hdr[c]) for c in range(COL_GRID_INI, COL_GRID_FIN) if isinstance(hdr[c], time)]

        registros, para_cargar = [], []
        for r in filas_xl[1:]:
            fecha = r[COL_FECHA]
            if r[COL_YEAR] != anio or not isinstance(fecha, datetime) or fecha.month != mes:
                continue
            trabajados = [t for c, t in grid if str(r[c]).strip() in ("1", "1.0")]
            trab = normalizar_trabajador(r[COL_AGENTE])
            dia = DIAS_ORDEN[fecha.weekday()]
            if trabajados:
                entrada, salida = trabajados[0], _salida(trabajados[-1])
                registros.append((fecha.date().isoformat(), dia, trab,
                                  entrada.strftime("%H:%M"), salida.strftime("%H:%M")))
            else:
                entrada = salida = None
                registros.append((fecha.date().isoformat(), dia, trab, "—", "—"))
            para_cargar.append({
                "semana_inicio": get_semana_inicio(fecha.date()), "trabajador": trab,
                "dia": dia, "entrada": entrada, "salida": salida,
                "es_libre": not trabajados,
            })

        if not registros:
            self.stdout.write(self.style.WARNING(f"Sin datos para {opts['mes']}."))
            return

        self.stdout.write(f"{opts['mes']}: {len(registros)} filas")
        for fecha_s, dia, trab, ent, sal in registros[:16]:
            self.stdout.write(f"  {fecha_s}  {dia:<10} {trab:<6} {ent}-{sal}")
        if len(registros) > 16:
            self.stdout.write(f"  ... (+{len(registros) - 16} filas)")

        if opts["cargar"]:
            n = guardar_turnos(para_cargar, on_ok=lambda c: None)
            self.stdout.write(self.style.SUCCESS(f"Cargados {n} turnos en TurnoEquipo."))
        else:
            self.stdout.write("(usa --cargar para insertarlos en la base de datos)")
