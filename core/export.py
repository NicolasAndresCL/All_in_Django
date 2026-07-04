"""
core/export.py — Generación de Excel (openpyxl) y PDF (fpdf2) desde filas genéricas.

Funciones puras que reciben columnas + filas (list[dict]) y devuelven `bytes`.
Las usan las acciones de exportación de la API.
"""

from io import BytesIO

from fpdf import FPDF


def generar_excel(columnas: list[str], filas: list[dict], hoja: str = "Datos") -> bytes:
    """Construye un .xlsx en memoria. Usa openpyxl directamente (sin pandas)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = hoja
    ws.append(columnas)
    for fila in filas:
        ws.append([fila.get(c, "") for c in columnas])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generar_pdf(titulo: str, columnas: list[str], filas: list[dict]) -> bytes:
    """Tabla simple en PDF (fpdf2). `pdf.output()` ya devuelve bytes."""
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, titulo, ln=True, align="C")
    pdf.ln(3)

    ancho = 190 // max(len(columnas), 1)
    pdf.set_font("Arial", "B", 9)
    pdf.set_fill_color(220, 220, 220)
    for col in columnas:
        pdf.cell(ancho, 8, str(col), border=1, align="C", fill=True)
    pdf.ln()

    pdf.set_font("Arial", "", 9)
    for fila in filas:
        for col in columnas:
            pdf.cell(ancho, 7, str(fila.get(col, "")), border=1)
        pdf.ln()
    return bytes(pdf.output())
